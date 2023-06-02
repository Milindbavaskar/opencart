import openpyxl


def Rowcount(file, sheetname):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return (sheet.max_row)


def ReadData(file, sheetname, rownum, colnum):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row=rownum, column=colnum).value


def WriteData(file, sheetname, rownum, colnum, data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    sheet.cell(row=rownum, column=colnum).value = data
    book.save(file)